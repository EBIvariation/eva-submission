# Copyright 2020 EMBL - European Bioinformatics Institute
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""
This module was borrow and modified from
https://github.com/EBIvariation/amp-t2d-submissions/blob/master/xls2xml/xls2xml/XLSReader.py

This module reads an Excel file and allows the user to get all the valid worksheet names,
get the 1st line in a worksheet and iterate over the rest of the worksheet row by row
(next_row). The returned row is a hash which contains only the keys that are defined in
a configuration file.
This module depends on openpyxl and pyyaml.
"""

import yaml
from ebi_eva_common_pyutils.logger import AppLogger
from openpyxl import load_workbook

WORKSHEETS_KEY_NAME = 'worksheets'
REQUIRED_HEADERS_KEY_NAME = 'required'
OPTIONAL_HEADERS_KEY_NAME = 'optional'
HEADERS_KEY_ROW = 'header_row'


class XLSBaseParser(AppLogger):
    """
    Base parser for Excel file for the fields from worksheets defined in a configuration file.
    It implements the base functioanlity allowing to open and validate the spreadsheet
    """

    def __init__(self, xls_filename, conf_filename, read_only=True):
        """
        Constructor
        :param xls_filename: Excel file path
        :type xls_filename: basestring
        :param conf_filename: configuration file path
        :type conf_filename: basestring
        """
        with open(conf_filename, 'r') as conf_file:
            self.xls_conf = yaml.full_load(conf_file)
        try:
            self.workbook = load_workbook(xls_filename, read_only=read_only)
        except Exception as e:
            self.error('Error loading %s', xls_filename)
            raise e
        self.worksheets = None
        self._active_worksheet = None
        self.row_offset = {}
        self.headers = {}
        self.valid = None

    @property
    def active_worksheet(self):
        return self._active_worksheet

    @active_worksheet.setter
    def active_worksheet(self, worksheet):
        if self.worksheets is None:
            self.valid_worksheets()
        if worksheet not in self.worksheets:
            raise ValueError('Worksheet ' + worksheet + ' is not valid!')

        self._active_worksheet = worksheet

    def valid_worksheets(self):
        """
        Get the list of the names of worksheets which have all the configured required headers
        :return: list of valid worksheet names in the Excel file
        :rtype: list
        """
        if self.worksheets is not None:
            return self.worksheets

        self.worksheets = []
        sheet_titles = self.workbook.sheetnames

        for title in self.xls_conf[WORKSHEETS_KEY_NAME]:
            # Check worksheet exists
            if title not in sheet_titles:
                continue

            # Check number of rows
            worksheet = self.workbook[title]
            header_row = self.xls_conf[title].get(HEADERS_KEY_ROW, 1)
            if worksheet.max_row < header_row + 1:
                continue
            # Check required headers are present
            self.headers[title] = [cell.value if cell.value is None else cell.value.strip()
                                   for cell in worksheet[header_row]]
            required_headers = self.xls_conf[title].get(REQUIRED_HEADERS_KEY_NAME, [])
            if set(required_headers) <= set(self.headers[title]):  # issubset
                self.worksheets.append(title)
            else:
                self.warning('Worksheet '+title+' does not have all the required headers!')
                self.valid = False

        return self.worksheets

    def get_valid_conf_keys(self):
        """
        :return: the list of valid worksheet names
        :rtype: list
        """
        return self.valid_worksheets()

    def is_valid(self):
        """
        Check that is all the worksheets contain required headers
        :return: True if all the worksheets contain required headers. False otherwise
        :rtype: bool
        """
        if self.valid is None:
            self.valid = True
            self.valid_worksheets()

        return self.valid


class XLSReader(XLSBaseParser):
    """
    Reader for Excel file for the fields from worksheets defined in a configuration file
    """

    def __init__(self, xls_filename, conf_filename):
        """
        Constructor
        :param xls_filename: Excel file path
        :type xls_filename: basestring
        :param conf_filename: configuration file path
        :type conf_filename: basestring
        """
        super().__init__(xls_filename, conf_filename, read_only=True)

    def __iter__(self):
        return self

    def base_row_offset(self, worksheet):
        return self.xls_conf[worksheet].get(HEADERS_KEY_ROW, 1)

    def next(self):
        """
        Retrieve next data row
        :return: A hash containing all the REQUIRED and OPTIONAL fields as keys
                and the corresponding data as values
        :rtype: dict
        """
        worksheet = self.active_worksheet
        if worksheet is None:
            self.warning('No worksheet is specified!')
            raise StopIteration

        if worksheet not in self.row_offset:
            self.row_offset[worksheet] = self.base_row_offset(worksheet)
        self.row_offset[worksheet] += 1

        required_headers = self.xls_conf[worksheet].get(REQUIRED_HEADERS_KEY_NAME, [])
        optional_headers = self.xls_conf[worksheet].get(OPTIONAL_HEADERS_KEY_NAME, [])

        for row in self.workbook[worksheet].iter_rows(min_row=self.row_offset[worksheet]):
            num_cells = 0
            for cell in row:
                num_cells += 1

            data = {}
            has_notnull = False
            for header in required_headers+optional_headers:
                header_index = num_cells
                if header in self.headers[worksheet]:
                    header_index = self.headers[worksheet].index(header)
                if header_index >= num_cells:
                    data[header] = None
                    continue

                cell = row[header_index]
                if cell.value is not None:
                    has_notnull = True

                data[header] = cell.value

            if has_notnull:
                data['row_num'] = self.row_offset[worksheet]
                return data

            # no data on this row, continue to next
            self.row_offset[worksheet] += 1

        raise StopIteration


class XLSWriter(XLSBaseParser):
    """
    Writer for Excel file for the fields from worksheets defined in a configuration file
    """

    def __init__(self, xls_filename, conf_filename):
        """
        Constructor
        :param xls_filename: Excel file path
        :type xls_filename: basestring
        :param conf_filename: configuration file path
        :type conf_filename: basestring
        """
        super().__init__(xls_filename, conf_filename, read_only=False)

    def edit_row(self, row_data: dict, remove_when_missing_values=True):
        worksheet = self.active_worksheet
        if worksheet is None:
            raise ValueError('No worksheet is specified!')

        if 'row_num' not in row_data:
            raise KeyError('No row specified in dict ' + str(row_data))
        row_num = row_data['row_num']

        required_headers = self.xls_conf[worksheet].get(REQUIRED_HEADERS_KEY_NAME, [])
        optional_headers = self.xls_conf[worksheet].get(OPTIONAL_HEADERS_KEY_NAME, [])

        for header in required_headers:
            header_index = self.headers[worksheet].index(header)
            if header not in row_data:
                raise ValueError('Header {0} is required but is not provided in row {1}'.format(header, row_num))
            self.workbook[worksheet].cell(column=header_index+1, row=row_num, value=row_data[header])

        for header in optional_headers:
            if header in self.headers[worksheet]:
                header_index = self.headers[worksheet].index(header)
                if header not in row_data and remove_when_missing_values:
                    # When data is missing remove the value from the cell
                    self.workbook[worksheet].cell(column=header_index+1, row=row_num, value='')
                elif header in row_data:
                    self.workbook[worksheet].cell(column=header_index+1, row=row_num, value=row_data[header])

    def set_rows(self, rows):
        """
        Write a set of rows from the top of the spreadsheet.
        """
        worksheet = self.active_worksheet
        if worksheet is None:
            raise ValueError('No worksheet is specified!')

        first_row = self.xls_conf[worksheet].get(HEADERS_KEY_ROW, 1) + 1
        for i, row in enumerate(rows):
            row['row_num'] = first_row + i
            self.edit_row(row)

    def save(self, filename):
        self.workbook.save(filename)
