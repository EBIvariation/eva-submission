# Copyright 2020 EMBL - European Bioinformatics Institute
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""
This module was borrow and modified from
https://github.com/EBIvariation/amp-t2d-submissions/blob/master/xls2xml/xls2xml/XLSReader.py

This module reads an Excel file and allows the user to get all the valid worksheet names,
get the 1st line in a worksheet and iterate over the rest of the worksheet row by row
(next_row). The returned row is a hash which contains only the keys that are defined in
a configuration file.
This module depends on openpyxl and pyyaml.
"""
import os
from collections import defaultdict

from cached_property import cached_property
from ebi_eva_common_pyutils.logger import AppLogger
from openpyxl import load_workbook
import yaml

WORKSHEETS_KEY_NAME = 'worksheets'
REQUIRED_HEADERS_KEY_NAME = 'required'
OPTIONAL_HEADERS_KEY_NAME = 'optional'
HEADERS_KEY_ROW = 'header_row'


class EVAXLSReader(AppLogger):

    def __init__(self, metadata_file):
        conf = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'etc', 'eva_project_conf.yaml')
        self.reader = XLSReader(metadata_file, conf)
        self.metadata_file=metadata_file

    def _get_all_rows(self, active_sheet):
        self.reader.active_worksheet = active_sheet
        rows = []
        try:
            r = self.reader.next()
            while r:
                rows.append(r)
                r = self.reader.next()
            rows.append(r)
        except StopIteration:
            pass
        return rows

    @cached_property
    def project(self):
        self.reader.active_worksheet = 'Project'
        try:
            return self.reader.next()
        except StopIteration:
            self.error('No project was found in the spreadsheet %s', self.metadata_file)

    @cached_property
    def analysis(self):
        return self._get_all_rows('Analysis')

    @cached_property
    def samples(self):
        return self._get_all_rows('Sample')

    @cached_property
    def files(self):
        return self._get_all_rows('Files')

    @cached_property
    def project_title(self):
        if self.project:
            return self.project.get('Project Title')

    @property
    def analysis_titles(self):
        return [a.get('Analysis Title') for a in self.analysis]

    @property
    def references(self):
        return list(set([a.get('Reference') for a in self.analysis if a.get('Reference')]))

    @property
    def samples_per_analysis(self):
        samples_per_analysis = defaultdict(list)
        for row in self.samples:
            samples_per_analysis[row.get('Analysis Alias')].append(row)
        return samples_per_analysis

    @property
    def files_per_analysis(self):
        files_per_analysis = defaultdict(list)
        for row in self.files:
            files_per_analysis[row.get('Analysis Alias')].append(row)
        return files_per_analysis


class XLSReader(AppLogger):
    """
    Reader for Excel file for the fields from worksheets defined in a configuration file
    """

    def __init__(self, xls_filename, conf_filename):
        """
        Constructor
        :param xls_filename: Excel file path
        :type xls_filename: basestring
        :param conf_filename: configuration file path
        :type conf_filename: basestring
        """
        with open(conf_filename, 'r') as conf_file:
            self.xls_conf = yaml.full_load(conf_file)
        try:
            self.workbook = load_workbook(xls_filename, read_only=True)
        except Exception as e:
            self.error('Error loading %s', xls_filename)
            raise e
        self.worksheets = None
        self._active_worksheet = None
        self.row_offset = {}
        self.headers = {}
        self.valid = None

    def __iter__(self):
        return self

    @property
    def active_worksheet(self):
        return self._active_worksheet

    @active_worksheet.setter
    def active_worksheet(self, worksheet):
        self._active_worksheet = worksheet

    def valid_worksheets(self):
        """
        Get the list of the names of worksheets which have all the configured required headers
        :return: list of valid worksheet names in the Excel file
        :rtype: list
        """
        if self.worksheets is not None:
            return self.worksheets

        self.worksheets = []
        sheet_titles = self.workbook.sheetnames

        for title in self.xls_conf[WORKSHEETS_KEY_NAME]:
            # Check worksheet exists
            if title not in sheet_titles:
                continue

            # Check number of rows
            worksheet = self.workbook[title]
            header_row = self.xls_conf[title].get(HEADERS_KEY_ROW, 1)
            if worksheet.max_row < header_row + 1:
                continue
            # Check required headers are present
            self.headers[title] = [cell.value if cell.value is None else cell.value.strip()
                                   for cell in worksheet[header_row]]
            required_headers = self.xls_conf[title].get(REQUIRED_HEADERS_KEY_NAME, [])
            if set(required_headers) <= set(self.headers[title]):  # issubset
                self.worksheets.append(title)
            else:
                self.warning('Worksheet '+title+' does not have all the required headers!')
                self.valid = False

        return self.worksheets

    def get_valid_conf_keys(self):
        """
        :return: the list of valid worksheet names
        :rtype: list
        """
        return self.valid_worksheets()

    def set_current_conf_key(self, current_key):
        """
        Set the active_worksheet with value in $current_key
        :param current_key: the name of the worksheet
        :type current_key:  basestring
        :return: nothing
        :rtype: void
        """
        self.active_worksheet = current_key

    def is_valid(self):
        """
        Check that is all the worksheets contain required headers
        :return: True if all the worksheets contain required headers. False otherwise
        :rtype: bool
        """
        if self.valid is None:
            self.valid = True
            self.valid_worksheets()

        return self.valid

    def get_current_headers(self):
        """
        Retrieve the list of worksheets that have all the required headers
        :return: the list of valid worksheet names in the Excel file
        :rtype: list
        """
        worksheets = self.valid_worksheets()
        current_worksheet = self.active_worksheet
        if current_worksheet not in worksheets:
            raise Exception('Worksheet '+current_worksheet+' is not available or not valid!')

        return [x for x in self.headers[current_worksheet] if x is not None]

    def next(self):
        """
        Retrieve next data row
        :param worksheet: the name of the worksheet
        :type worksheet: basestring
        :return: A hash containing all the REQUIRED and OPTIONAL fields as keys
                and the corresponding data as values
        :rtype: dict
        """
        if self.worksheets is None:
            self.valid_worksheets()

        worksheet = self.active_worksheet
        if worksheet is None:
            self.warning('No worksheet is specified!')
            raise StopIteration
        if worksheet not in self.worksheets:
            self.warning('Worksheet ' + worksheet + ' is not valid!')
            raise StopIteration

        if worksheet not in self.row_offset:
            self.row_offset[worksheet] = self.xls_conf[worksheet].get(HEADERS_KEY_ROW, 1)
        self.row_offset[worksheet] += 1

        required_headers = self.xls_conf[worksheet].get(REQUIRED_HEADERS_KEY_NAME, [])
        optional_headers = self.xls_conf[worksheet].get(OPTIONAL_HEADERS_KEY_NAME, [])

        for row in self.workbook[worksheet].iter_rows(min_row=self.row_offset[worksheet]):
            num_cells = 0
            for cell in row:
                num_cells += 1

            data = {}
            has_notnull = False
            for header in required_headers+optional_headers:
                header_index = num_cells
                if header in self.headers[worksheet]:
                    header_index = self.headers[worksheet].index(header)
                if header_index >= num_cells:
                    data[header] = None
                    continue

                cell = row[header_index]
                if cell.value is not None:
                    has_notnull = True

                data[header] = cell.value

            if has_notnull:
                data['row_num'] = self.row_offset[worksheet]
                return data

            # no data on this row, continue to next
            self.row_offset[worksheet] += 1

        raise StopIteration