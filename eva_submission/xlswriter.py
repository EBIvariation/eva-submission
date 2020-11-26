# Copyright 2020 EMBL - European Bioinformatics Institute
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import os
from collections import defaultdict

from cached_property import cached_property
from ebi_eva_common_pyutils.logger import AppLogger
from openpyxl import load_workbook
import yaml

from eva_submission.xlsreader import XLSReader, REQUIRED_HEADERS_KEY_NAME, OPTIONAL_HEADERS_KEY_NAME, HEADERS_KEY_ROW


class XLSWriter(XLSReader):
    """
    Writer for Excel file for the fields from worksheets defined in a configuration file
    """

    def __init__(self, xls_filename, conf_filename):
        """
        Constructor
        :param xls_filename: Excel file path
        :type xls_filename: basestring
        :param conf_filename: configuration file path
        :type conf_filename: basestring
        """
        XLSReader.__init__(self, xls_filename, conf_filename)

        try:
            self.workbook = load_workbook(xls_filename)
        except Exception as e:
            self.error('Error loading %s', xls_filename)
            raise e

    def edit_row(self, row_data: dict, remove_when_missing_values=True):
        worksheet = self.active_worksheet
        if worksheet is None:
            raise ValueError('No worksheet is specified!')

        if 'row_num' not in row_data:
            raise KeyError('No row specified in dict ' + str(row_data))
        row_num = row_data['row_num']

        required_headers = self.xls_conf[worksheet].get(REQUIRED_HEADERS_KEY_NAME, [])
        optional_headers = self.xls_conf[worksheet].get(OPTIONAL_HEADERS_KEY_NAME, [])

        for header in required_headers:
            header_index = self.headers[worksheet].index(header)
            if header not in row_data:
                raise ValueError('Header {0} is required but is not provided in row {1}'.format(header, row_num))
            self.workbook[worksheet].cell(column=header_index+1, row=row_num, value=row_data[header])

        for header in optional_headers:
            if header in self.headers[worksheet]:
                header_index = self.headers[worksheet].index(header)
                if header not in row_data and remove_when_missing_values:
                    # When data is missing remove the value from the cell
                    self.workbook[worksheet].cell(column=header_index+1, row=row_num, value='')
                elif header in row_data:
                    self.workbook[worksheet].cell(column=header_index+1, row=row_num, value=row_data[header])

    def set_rows(self, rows):
        """
        Write a set of rows from the top of the spreadsheet.
        """
        worksheet = self.active_worksheet
        if worksheet is None:
            raise ValueError('No worksheet is specified!')

        first_row = self.xls_conf[worksheet].get(HEADERS_KEY_ROW, 1) + 1
        for i, row in enumerate(rows):
            row['row_num'] = first_row + i
            self.edit_row(row)

    def save(self, filename):
        self.workbook.save(filename)


class EVAXLSWriter(AppLogger):

    def __init__(self, metadata_source, metadata_dest=None):
        conf = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'etc', 'eva_project_conf.yaml')
        self.writer = XLSWriter(metadata_source, conf)
        self.metadata_source = metadata_source
        if metadata_dest:
            self.metadata_dest = metadata_dest
        else:
            self.metadata_dest = metadata_source

    def _set_all_rows(self, active_sheet, rows):
        self.writer.active_worksheet = active_sheet
        self.writer.set_rows(rows)

    def save(self):
        self.writer.save(self.metadata_dest)

    def set_files(self, file_dicts):
        self._set_all_rows('Files', file_dicts)

    def set_project(self, project_dict):
        self._set_all_rows('Project', [project_dict])

    def set_analysis(self, analysis_dicts):
        self._set_all_rows('Analysis', analysis_dicts)

    def set_samples(self, sample_dicts):
        self._set_all_rows('Sample', sample_dicts)

    def set_files(self, file_dicts):
        self._set_all_rows('Files', file_dicts)

