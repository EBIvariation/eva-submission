import os
from unittest import TestCase
from unittest.mock import patch

import openpyxl

from eva_sub_cli_processing.sub_cli_to_eload_converter.json_to_xlsx_converter import JsonToXlsxConverter
from eva_submission import ROOT_DIR
from eva_submission.xlsx.xlsx_validation import EvaXlsxValidator


class TestJsonToXlsxConverter(TestCase):
    def setUp(self) -> None:
        self.resources_folder = os.path.join(ROOT_DIR, 'tests', 'resources')
        self.test_input_json = os.path.join(self.resources_folder, 'input_json_for_json_to_xlsx_converter.json')
        self.output_xlsx = os.path.join(self.resources_folder, 'json_to_xlsx_converted.xlsx')

    def tearDown(self):
        files_from_tests = [
            os.path.join(self.resources_folder, 'json_to_xlsx_converted.xlsx')
        ]
        for f in files_from_tests:
            if os.path.exists(f):
                os.remove(f)

    def test_json_to_xlsx_converter(self):
        self.converter = JsonToXlsxConverter(self.test_input_json, self.output_xlsx)

        self.converter.convert_json_to_xlsx()
        output_xlsx_data = self.read_excel_to_dict(self.output_xlsx)

        self.assert_submitter_details(output_xlsx_data['Submitter Details'])
        self.assert_project_details(output_xlsx_data['Project'])
        self.assert_analysis_details(output_xlsx_data['Analysis'])
        self.assert_samples_details(output_xlsx_data['Sample'])
        self.assert_files_details(output_xlsx_data['Files'])

    def assert_submitter_details(self, submitter_details):
        assert len(submitter_details) == 2
        assert submitter_details[0]['First Name'] == 'John'
        assert submitter_details[0]['Last Name'] == 'Smith'
        assert submitter_details[0]['Email Address'] == 'john.smith@example.com'
        assert submitter_details[0]['Center'] == 'University of Example'
        assert submitter_details[0]['Laboratory'] == 'Genomics Lab'
        assert submitter_details[0]['Address'] == '1 street address'

        assert submitter_details[1]['First Name'] == 'Jane'
        assert submitter_details[1]['Last Name'] == 'Doe'
        assert submitter_details[1]['Email Address'] == 'jane.doe@example.com'
        assert submitter_details[1]['Center'] == 'University of Example'
        assert submitter_details[1]['Laboratory'] == 'Bioinformatics Lab'
        assert submitter_details[1]['Address'] == '1 street address'

    def assert_project_details(self, project_details):
        assert len(project_details) == 1
        assert project_details[0]['Project Title'] == 'Example Project'
        assert project_details[0]['Project Alias'] == 'Example Project'
        assert project_details[0]['Description'] == 'An example project for demonstration purposes'
        assert project_details[0]['Center'] == 'University of Example'
        assert project_details[0]['Tax ID'] == 9606
        assert project_details[0]['Parent Project'] == 'PRJEB00001'
        assert project_details[0]['Child Project(s)'] == 'PRJEB00002,PRJEB00003'
        assert project_details[0]['Hold Date'].strftime('%Y-%m-%d') == '2023-12-31'

    def assert_analysis_details(self, analysis_details):
        assert len(analysis_details) == 3
        assert analysis_details[0]['Analysis Alias'] == 'VD1'
        assert analysis_details[0]['Analysis Title'] == 'Variant Detection 1'
        assert analysis_details[0]['Description'] == 'An example analysis for demonstration purposes'
        assert analysis_details[0]['Project Title'] == 'Example Project'
        assert analysis_details[0]['Experiment Type'] == 'Whole genome sequencing'
        assert analysis_details[0]['Imputation'] == '1'
        assert analysis_details[0]['Platform'] == 'BGISEQ-500'
        assert analysis_details[0]['Reference Fasta Path'] == 'GCA_000001405.27_fasta.fa'
        assert analysis_details[0]['Reference'] == 'GCA_000001405.27'

        assert analysis_details[1]['Analysis Alias'] == 'VD2'
        assert analysis_details[1]['Analysis Title'] == 'Variant Detection 2'
        assert analysis_details[1]['Description'] == 'An example analysis for demonstration purposes'
        assert analysis_details[0]['Project Title'] == 'Example Project'
        assert analysis_details[1]['Experiment Type'] == 'Whole genome sequencing'
        assert analysis_details[1]['Platform'] == 'BGISEQ-500'
        assert analysis_details[1]['Reference Fasta Path'] == 'GCA_000001405.27_fasta.fa'
        assert analysis_details[1]['Reference'] == 'GCA_000001405.27'

        assert analysis_details[2]['Analysis Alias'] == 'VD3'
        assert analysis_details[2]['Analysis Title'] == 'Variant Detection 3'
        assert analysis_details[2]['Description'] == 'An example analysis for demonstration purposes'
        assert analysis_details[0]['Project Title'] == 'Example Project'
        assert analysis_details[2]['Experiment Type'] == 'Whole genome sequencing'
        assert analysis_details[2]['Platform'] == 'BGISEQ-500'
        assert analysis_details[2]['Reference Fasta Path'] == 'GCA_000001405.27_fasta.fa'
        assert analysis_details[2]['Reference'] == 'GCA_000001405.27'

    def assert_samples_details(self, samples_details):
        assert len(samples_details) == 4
        assert samples_details[0]['Analysis Alias'] == 'VD1,VD2,VD3'
        assert samples_details[0]['Sample Accession'] == 'SAME00001'
        assert samples_details[0]['Sample ID'] == 'sample1'

        assert samples_details[1]['Analysis Alias'] == 'VD1,VD2,VD3'
        assert samples_details[1]['Sample Accession'] == 'SAME00002'
        assert samples_details[1]['Sample ID'] == 'sample2'

        assert samples_details[2]['Analysis Alias'] == 'VD3'
        assert samples_details[2]['Sample Accession'] == 'SAME00003'
        assert samples_details[2]['Sample ID'] == 'sample3'

        assert samples_details[3]['Analysis Alias'] == 'VD4,VD5'
        assert samples_details[3]['Sample Name'] == 'Lm_17_S8'
        assert samples_details[3]['Title'] == 'Bastet normal sample'
        assert samples_details[3]['collection_date'].strftime('%Y-%m-%d') == '2021-03-12'
        assert samples_details[3]['Description'] == 'Test Description'
        assert samples_details[3]['Sample ID'] == 'sample4'
        assert samples_details[3]['Scientific Name'] == 'Lemur catta'
        assert samples_details[3]['sex'] == 'Female'
        assert samples_details[3]['Tax Id'] == 9447
        assert samples_details[3]['tissue_type'] == 'skin'

    def assert_files_details(self, files_details):
        assert len(files_details) == 3
        assert files_details[0]['Analysis Alias'] == 'VD1'
        assert files_details[0]['File Name'] == 'example1.vcf.gz'
        assert files_details[0]['File Type'] == 'vcf'
        assert files_details[1]['Analysis Alias'] == 'VD2'
        assert files_details[1]['File Name'] == 'example2.vcf'
        assert files_details[1]['File Type'] == 'vcf'
        assert files_details[2]['Analysis Alias'] == 'VD3'
        assert files_details[2]['File Name'] == 'example3.vcf'
        assert files_details[2]['File Type'] == 'vcf'

    def read_excel_to_dict(self, file_path):
        wb = openpyxl.load_workbook(file_path, data_only=True)
        data_dict = {}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            header_row = 3 if sheet_name == "Sample" else 1
            data_start_row = header_row + 1

            header_cells = next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True), None)
            if not header_cells:
                continue

            headers = [cell for cell in header_cells if cell is not None]

            data = []
            for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
                row_dict = {key: value for key, value in zip(headers, row) if
                            key is not None and value is not None}
                if row_dict:
                    data.append(row_dict)

            data_dict[sheet_name] = data

        return data_dict

    def test_json_to_xlsx_converter_is_valid(self):
        self.converter = JsonToXlsxConverter(self.test_input_json, self.output_xlsx)

        self.converter.convert_json_to_xlsx()
        with (
            patch.object(EvaXlsxValidator, 'check_biosamples_accessions'),
            patch('eva_submission.xlsx.xlsx_validation.check_existing_project_in_ena')
        ):
            validator = EvaXlsxValidator(self.output_xlsx)
            validator.validate()
            assert validator.error_list == ['Check Analysis Alias vs Samples: VD4,VD5 present in Samples not in Analysis Alias']
