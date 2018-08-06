from openpyxl import load_workbook
import os
import argparse
from collections import OrderedDict
import sys

ORIGINAL_SAMPLE_SHEET = "Sample"
ORIGINAL_FILES_SHEET = "Files"
SAMPLE_NAME_FIELD = "Sample Name"
SAMPLE_ID_FIELD = "Sample ID"
FILE_NAME_FIELD = "File Name"
FILE_TYPE_FIELD = "File Type"
REWRITTEN_SAMPLE_NAMES_SHEET = "Sample_Names"
REWRITTEN_FILE_NAMES_SHEET = "File_Names"


def cell_value_empty(cell_value):
    return cell_value == "None" or cell_value == ""


def find_cell_coords_with_text(eva_sample_sheet, num_rows, num_cols, text_to_find):
    for i in range(1, num_rows + 1):
        for j in range(1, num_cols + 1):
            if str(eva_sample_sheet.cell(None, i, j).value).strip().lower() == text_to_find.lower():
                return i, j
    raise Exception("ERROR: Could not find cell with text '{0}' in the Sample tab!".format(text_to_find))


def get_sample_names(eva_sample_sheet):
    """
    Get sample names from either the Novel Sample section or the Pre-registered sample section
    """
    sample_names = []
    num_rows = eva_sample_sheet.max_row
    num_cols = eva_sample_sheet.max_column

    i, j = find_cell_coords_with_text(eva_sample_sheet, num_rows, num_cols, SAMPLE_NAME_FIELD)

    prereg_sample_id_col_present = \
        str(eva_sample_sheet.cell(None, i, j - 4).value).strip().lower() == SAMPLE_ID_FIELD.lower()
    while i <= num_rows:
        i += 1
        # In the previous version of the template, Pre-registered sample section doesn't have Sample ID
        # In such cases, we can only use Sample names from the Novel sample section
        if not prereg_sample_id_col_present:
            sample_name_from_prereg_section = ""
        else:
            sample_name_from_prereg_section = str(eva_sample_sheet.cell(None, i, j - 4).value).strip()

        sample_name_from_novel_section = str(eva_sample_sheet.cell(None, i, j).value).strip()
        if not cell_value_empty(sample_name_from_prereg_section) \
                and not cell_value_empty(sample_name_from_novel_section):
            raise Exception("ERROR: Both Novel Sample Names and Pre-registered sample names are present "
                            "in the Metadata sheet. Only one of these should be present!")
        if cell_value_empty(sample_name_from_prereg_section) \
                and cell_value_empty(sample_name_from_novel_section):
            continue
        sample_name = sample_name_from_novel_section if cell_value_empty(sample_name_from_prereg_section) \
            else sample_name_from_prereg_section
        sample_names.append(sample_name)
    return sample_names


def get_file_names(eva_files_sheet):
    """
    Get file names from either the Novel Sample section or the Pre-registered sample section
    """
    file_names = OrderedDict()
    num_rows = eva_files_sheet.max_row
    num_cols = eva_files_sheet.max_column

    i, j = find_cell_coords_with_text(eva_files_sheet, num_rows, num_cols, "File Name")
    while i <= num_rows:
        i += 1
        file_name = os.path.basename(str(eva_files_sheet.cell(None, i, j).value).strip())
        file_type = str(eva_files_sheet.cell(None, i, j + 1).value).strip()
        if cell_value_empty(file_name):
            continue
        file_names[file_name] = file_type
    return file_names


def rewrite_samples_tab(metadata_file_copy, sample_names):
    """
    Re-write the Sample names from the Samples tab:\n
    Since the samples_checker utility expects data in a single column with header, re-write the sample names
    in this expected format to a separate tab "Sample_Names"
    """
    if REWRITTEN_SAMPLE_NAMES_SHEET not in metadata_file_copy.sheetnames:
        sample_name_sheet = metadata_file_copy.create_sheet(REWRITTEN_SAMPLE_NAMES_SHEET)
    else:
        sample_name_sheet = metadata_file_copy[REWRITTEN_SAMPLE_NAMES_SHEET]
    sample_name_sheet.cell(None, 1, 1).value = SAMPLE_NAME_FIELD
    row_index = 2
    for sample_name in sample_names:
        sample_name_sheet.cell(None, row_index, 1).value = sample_name
        row_index += 1


def rewrite_files_tab(metadata_file_copy, file_name_types):
    """
    Re-write the File names from the Files tab:\n
    Since the samples_checker utility expects file names and not the ELOAD FTP paths, re-write the File names
    in this expected format to a separate tab "File_Names"
    """
    if REWRITTEN_FILE_NAMES_SHEET not in metadata_file_copy.sheetnames:
        file_name_sheet = metadata_file_copy.create_sheet(REWRITTEN_FILE_NAMES_SHEET)
    else:
        file_name_sheet = metadata_file_copy[REWRITTEN_FILE_NAMES_SHEET]
    file_name_sheet.cell(None, 1, 1).value = FILE_NAME_FIELD
    file_name_sheet.cell(None, 1, 2).value = FILE_TYPE_FIELD
    row_index = 2
    for file_name in file_name_types:
        file_name_sheet.cell(None, row_index, 1).value = file_name
        file_name_sheet.cell(None, row_index, 2).value = file_name_types[file_name]
        row_index += 1


def rewrite_samples_and_files_tab(eva_metadata_file):
    metadata_wb = load_workbook(eva_metadata_file, data_only=True)

    if ORIGINAL_SAMPLE_SHEET not in metadata_wb.sheetnames:
        raise Exception("{0} tab could not be found in the EVA metadata sheet: {1}".format(ORIGINAL_SAMPLE_SHEET,
                                                                                           eva_metadata_file))
    if ORIGINAL_FILES_SHEET not in metadata_wb.sheetnames:
        raise Exception("{0} tab could not be found in the EVA metadata sheet: {1}".format(ORIGINAL_FILES_SHEET,
                                                                                           eva_metadata_file))

    sample_names = get_sample_names(metadata_wb[ORIGINAL_SAMPLE_SHEET])
    file_name_types = get_file_names(metadata_wb[ORIGINAL_FILES_SHEET])

    eva_metadata_sheet_copy = os.path.dirname(os.path.realpath(__file__)) + os.path.sep + \
                              ".".join(os.path.basename(eva_metadata_file).split(".")[:-1]) + \
                              "_with_sample_names_file_names.xlsx"
    metadata_wb.save(eva_metadata_sheet_copy)
    metadata_wb.close()
    metadata_wb_copy = load_workbook(eva_metadata_sheet_copy)

    rewrite_samples_tab(metadata_wb_copy, sample_names)
    rewrite_files_tab(metadata_wb_copy, file_name_types)

    metadata_wb_copy.save(eva_metadata_sheet_copy)
    metadata_wb_copy.close()

    return eva_metadata_sheet_copy


def main():
    arg_parser = argparse.ArgumentParser(
        description='Transform and output validated data from an excel file to a XML file')
    arg_parser.add_argument('--samples-checker-dir', required=True, dest='samples_checker_dir',
                            help='Path to the directory with the Samples Checker module')
    arg_parser.add_argument('--xls2xml-dir', required=True, dest='xls2xml_dir',
                            help='Path to the directory with the xls2xml module')
    arg_parser.add_argument('--metadata-file', required=True, dest='metadata_file',
                            help='EVA Submission Metadata Excel sheet')
    arg_parser.add_argument('--vcf-files-path', required=True, dest='vcf_files_path',
                            help='Path to the directory in which submitted files can be found')

    args = arg_parser.parse_args()
    vcf_files_path = args.vcf_files_path
    metadata_file = args.metadata_file

    sys.path.append(args.samples_checker_dir)
    sys.path.append(args.xls2xml_dir)
    from xls2xml import xls2xml
    from samples_checker import check_samples

    data_dir = os.path.dirname(os.path.realpath(__file__))
    xls_conf = data_dir + os.path.sep + "conf/EVA_xls2xml_v2.conf"
    xls_schema = data_dir + os.path.sep + "conf/EVA_xls2xml_v2.schema"
    xslt_filename = data_dir + os.path.sep + "conf/EVA_xls2xml_v2.xslt"

    file_xml = os.path.splitext(metadata_file)[0] + ".file.xml"
    sample_xml = os.path.splitext(metadata_file)[0] + ".sample.xml"

    rewritten_metadata_file = rewrite_samples_and_files_tab(metadata_file)
    xls2xml.convert_xls_to_xml(xls_conf, [REWRITTEN_FILE_NAMES_SHEET], xls_schema, xslt_filename,
                               rewritten_metadata_file, file_xml)
    xls2xml.convert_xls_to_xml(xls_conf, [REWRITTEN_SAMPLE_NAMES_SHEET], xls_schema, xslt_filename,
                               rewritten_metadata_file, sample_xml)
    check_samples.get_sample_diff(vcf_files_path, file_xml, sample_xml)


if __name__ == "__main__":
    main()
