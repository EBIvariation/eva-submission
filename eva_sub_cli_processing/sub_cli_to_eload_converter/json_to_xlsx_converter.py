import json
from datetime import datetime

import openpyxl
from ebi_eva_common_pyutils.logger import AppLogger
from openpyxl.workbook import Workbook

def parse_date(x: str):
    try:
        return datetime.strptime(x, "%Y-%m-%d").date()
    except ValueError:
        pass
    try:
        return datetime.strptime(x, "%Y-%m").date()
    except ValueError:
        pass
    try:
        return datetime.strptime(x, "%Y").date()
    except ValueError:
        pass
    return x


# The dict contains the mapping between the field in JSON and the header of the Excel spreadsheet.
# when the value is a dict it must contain a 'name' to define the header and optionally can contain
#  - transform which define a function applied to the cell value
#  - default which define a default value when the key is not present in the JSON
#  - link which defines another value in the input JSON to apply in the format <worksheet>.<column> (works on project value) or <>.<column> to get data from the current row
json_to_xlsx_key_mapper = {
    "worksheets": {
        "submitterDetails": "Submitter Details",
        "project": "Project",
        "analysis": "Analysis",
        "sample": "Sample",
        "files": "Files"
    },

    "submitterDetails": {
        "lastName": "Last Name",
        "firstName": "First Name",
        "email": "Email Address",
        "laboratory": "Laboratory",
        "centre": "Center",
        "address": "Address",
        'DUMMY1': 'Telephone Number'
    },

    "project": {
        "title": "Project Title",
        "description": "Description",
        "centre": "Center",
        "taxId": {'name': "Tax ID", 'transform': lambda x: int(x)},
        "publications": {'name': "Publication(s)", 'transform': lambda x: ','.join(x)},
        "parentProject": "Parent Project",
        "childProjects": {'name': "Child Project(s)", 'transform': lambda x: ','.join(x)},
        "peerProjects": {'name': "Peer Project(s)", 'transform': lambda x: ','.join(x)},
        "links": {'name': "Link(s)", 'transform': lambda x: ','.join(x)},
        "holdDate": {'name': "Hold Date", 'transform': lambda x: datetime.strptime(x, "%Y-%m-%d").date()},
        "collaborators": "Collaborator(s)",
        "strain": "Strain",
        "breed": "Breed",
        "broker": "Broker",
        'DUMMY1': {'name':'Project Alias', 'link': 'project.title'}
    },

    "analysis": {
        "analysisTitle": "Analysis Title",
        "analysisAlias": "Analysis Alias",
        "description": "Description",
        "experimentType": "Experiment Type",
        "referenceGenome": "Reference",
        "referenceFasta": "Reference Fasta Path",
        "platform": "Platform",
        "software": {'name': "Software", 'transform': lambda x: ','.join(x)},
        "pipelineDescriptions": "Pipeline Description",
        "imputation": {'name': "Imputation", 'transform': lambda x: '1' if x == True else ''},
        "phasing": {'name': "Phasing", 'transform': lambda x: '1' if x == True else ''},
        "centre": "Centre",
        "date": {'name': "Date", 'transform': lambda x: datetime.strptime(x, "%Y-%m-%d").date()},
        "links": {'name': "Link(s)", 'transform': lambda x: ','.join(x)},
        "runAccessions": {'name': "Run Accession(s)", 'transform': lambda x: ','.join(x)},
        'Dummy': {'name': "Project Title", 'link': 'project.title'}
    },

    "sample": {
        "analysisAlias": {'name': "Analysis Alias", 'transform': lambda x: ','.join(x)},
        "sampleInVCF": "Sample ID",
        "bioSampleAccession": "Sample Accession",
        "bioSampleName": "Sample Name",
        "title": "Title",
        "description": "Description",
        "uniqueName": "Unique Name",
        "prefix": "Prefix",
        "subject": "Subject",
        "derivedFrom": "Derived From",
        "taxId": {'name': "Tax Id", 'transform': lambda x: int(x)},
        "scientificName": "Scientific Name",
        "commonName": "Common Name",
        "matingType": "mating_type",
        "sex": "sex",
        "population": "population",
        "cellType": "cell_type",
        "devStage": "dev_stage",
        "germline": "germline",
        "tissueLib": "tissue_lib",
        "tissueType": "tissue_type",
        "BioMaterial": "bio_material",
        "cultureCollection": "culture_collection",
        "specimenVoucher": "specimen_voucher",
        "collectedBy": "collected_by",
        "collectionDate": {'name': "collection_date", 'transform': parse_date},
        "geographicLocationCountrySea": "geographic location (country and/or sea)",
        "geographicLocationRegion": "geographic location (region and locality)",
        "host": "host",
        "identifiedBy": "identified_by",
        "isolationSource": "isolation_source",
        "latLon": "lat_lon",
        "LabHost": "lab_host",
        "environmentalSample": "environmental_sample",
        "cultivar": "cultivar",
        "ecotype": "ecotype",
        "isolate": "isolate",
        "strain": "strain",
        "subSpecies": "sub_species",
        "variety": "variety",
        "subStrain": "sub_strain",
        "cellLine": "cell_line",
        "serotype": "serotype",
        "serovar": "serovar"
    },

    "files": {
        "analysisAlias": "Analysis Alias",
        "fileName": "File Name",
        'DUMMY1': 'MD5',
        'DUMMY2': {'name': 'File Type', 'link': '.fileName', 'transform': lambda x: 'vcf' if x.endswith('.vcf.gz') or x.endswith('.vcf') else ''},
    }
}

class JsonToXlsxConverter(AppLogger):

    def __init__(self,  input_json_file, output_xlsx_file):
        self.input_json_file = input_json_file
        self.output_xlsx_file = output_xlsx_file
        with open(input_json_file, 'r') as f:
            self.data = json.load(f)

    def convert_json_to_xlsx(self):
        workbook = Workbook()
        self.create_dummy_instructions_sheet(workbook)
        for worksheet_key in json_to_xlsx_key_mapper['worksheets']:
            worksheet_title = json_to_xlsx_key_mapper['worksheets'][worksheet_key]
            worksheet_data = self.data[worksheet_key]
            self.create_worksheet(workbook, worksheet_key, worksheet_title, worksheet_data)
        workbook.save(self.output_xlsx_file)

    def create_worksheet(self, workbook, worksheet_key, worksheet_title, worksheet_data):
        self.create_worksheet_header(workbook, worksheet_key, worksheet_title)
        if worksheet_key == "submitterDetails":
            self._fill_in_worksheet(workbook, worksheet_title, worksheet_key, source_data_list=worksheet_data, start_row_index = 1)
        elif worksheet_key == "project":
            self._fill_in_worksheet(workbook, worksheet_title, worksheet_key, source_data_list=[worksheet_data], start_row_index=2)
        elif worksheet_key == "analysis":
            self._fill_in_worksheet(workbook, worksheet_title, worksheet_key, source_data_list=worksheet_data, start_row_index=1)
        elif worksheet_key == "sample":
            sample_flattened_data = self._flatten_sample_data(worksheet_data)
            self._fill_in_worksheet(workbook, worksheet_title, worksheet_key, source_data_list=sample_flattened_data, start_row_index=3)
        elif worksheet_key == "files":
            self._fill_in_worksheet(workbook, worksheet_title, worksheet_key, source_data_list=worksheet_data, start_row_index=1)

    def create_worksheet_header(self, workbook, worksheet_key, worksheet_title):
        workbook.create_sheet(worksheet_title)
        if worksheet_key == "sample":
            header_row_index = 3
        else:
            header_row_index = 1
        header_col_index = 0
        for header_key in sorted(json_to_xlsx_key_mapper[worksheet_key]):
            header_col_index += 1
            if isinstance(json_to_xlsx_key_mapper[worksheet_key][header_key], dict):
                header_value = json_to_xlsx_key_mapper[worksheet_key][header_key]['name']
            else:
                header_value = json_to_xlsx_key_mapper[worksheet_key][header_key]
            cell = workbook[worksheet_title].cell(column=header_col_index, row=header_row_index,
                                                  value=header_value)
            cell.font = openpyxl.styles.Font(bold=True)

    def create_dummy_instructions_sheet(self, workbook):
        worksheet_title = 'PLEASE READ FIRST'
        workbook.create_sheet(worksheet_title)
        workbook[worksheet_title].cell(column=1, row=3, value='V1.1.4 August 2020')

    def _fill_in_worksheet(self, output_workbook, output_worksheet_title, worksheet_key, source_data_list,  start_row_index=0):
        row_index = start_row_index
        for source_data_element in source_data_list:
            row_index += 1
            col_index = 0
            for header_key in sorted(json_to_xlsx_key_mapper[worksheet_key]):
                col_index += 1
                if header_key not in source_data_element:
                    cell_value = ''
                    if isinstance(json_to_xlsx_key_mapper[worksheet_key][header_key], dict):
                        header_dict = json_to_xlsx_key_mapper[worksheet_key][header_key]
                        if 'default' in header_dict and not cell_value:
                            cell_value = json_to_xlsx_key_mapper[worksheet_key][header_key].get('default')
                        if 'link' in header_dict and not cell_value:
                            tmp_worksheet, tmp_header = json_to_xlsx_key_mapper[worksheet_key][header_key].get('link').split('.')
                            if tmp_worksheet:
                                cell_value = self.data[tmp_worksheet][tmp_header]
                            else:
                                cell_value = source_data_element.get(tmp_header)
                else:
                    cell_value = source_data_element.get(header_key)
                if isinstance(json_to_xlsx_key_mapper[worksheet_key][header_key], dict):
                    header_dict = json_to_xlsx_key_mapper[worksheet_key][header_key]
                    if 'transform' in header_dict and cell_value:
                        cell_value = json_to_xlsx_key_mapper[worksheet_key][header_key].get('transform')(cell_value)

                output_workbook[output_worksheet_title].cell(column=col_index, row=row_index, value=cell_value)

    def _flatten_sample_data(self, sample_data):
        sample_flattened_data = []
        for sample in sample_data:
            sample_flattened_data.append(
                {
                    **sample,
                    **sample.get('bioSampleObject', {}),
                    **({'bioSampleName': sample['bioSampleObject']['name']}
                       if 'bioSampleObject' in sample and 'name' in sample['bioSampleObject'] and
                           sample['bioSampleObject']['name']
                        else {}),
                    **{key: char_data[0].get('text')
                        for key, char_data in
                        sample.get('bioSampleObject', {}).get('characteristics', {}).items()
                       }
                 }
            )
        # If both bioSampleName and sampleInVcf are both present and not equal, swap the values & warn.
        # This performs the sample check correctly, but these samples will be brokered using sampleInVcf
        # rather than bioSampleName.
        ignored_biosample_names = []
        for sample in sample_flattened_data:
            sampleInVcf = sample.get('sampleInVCF')
            bioSampleName = sample.get('bioSampleName')
            if sampleInVcf and bioSampleName and sampleInVcf != bioSampleName:
                sample['sampleInVCF'] = bioSampleName
                sample['bioSampleName'] = sampleInVcf
                ignored_biosample_names.append(bioSampleName)
        if ignored_biosample_names:
            self.warning(f'For some samples, BioSample names do not match sample name in VCF. '
                         f'Will ignore the following BioSample names: {",".join(ignored_biosample_names)}')
        return sample_flattened_data
